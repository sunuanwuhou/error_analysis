from __future__ import annotations

import argparse
import base64
import json
import posixpath
import re
import subprocess
import zipfile
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Callable, DefaultDict, Dict, Iterable, List, Sequence, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

SHEET_YANYU = '\u8a00\u8bed'
TYPE_YANYU = '\u8a00\u8bed\u7406\u89e3\u4e0e\u8868\u8fbe'
UNCATEGORIZED = '\u672a\u5206\u7c7b'
IMAGE_PROMPT_FALLBACK = '\u539f\u9898\u89c1\u56fe\u7247'

QUESTION_HEADER_ALIASES = {
    '\u9898\u578b': 'type',
    '\u5177\u4f53\u9898\u578b': 'subtype',
    '\u9898\u53f7/\u5957\u5377': 'paper_ref',
    '\u9898\u53f7': 'paper_ref',
    '\u9898\u76ee': 'question',
    '\u6b63\u786e\u7b54\u6848': 'answer',
    '\u9519\u56e0\u5206\u7c7b': 'error_reason',
    '\u9519\u8bef\u539f\u56e0': 'error_reason',
    '\u9519\u8bef\u5177\u4f53\u539f\u56e0': 'root_reason',
    '\u9519\u8bef\u5206\u6790': 'root_reason',
    '\u91cd\u8981\u603b\u7ed3': 'analysis',
    '\u4e8c\u5237\u60c5\u51b5': 'review_status',
    '\u5728\u5237\u60c5\u51b5': 'review_status',
    '\u5907\u6ce8': 'review_status',
    '\u540c\u7c7b\u9898\u76ee': 'similar_questions',
    '\u7c7b\u4f3c\u9898\u76ee': 'similar_questions',
}

NOTE_KEYWORDS = (
    '\u5927\u7eb2',
    '\u6982\u62ec\u603b\u7ed3',
    '\u6574\u4f53\u505a\u9898\u601d\u8def',
    '\u5e38\u89c1\u603b\u7ed3',
    '\u5751\u70b9',
    '\u6ce8\u610f\u70b9',
    '\u505a\u9898\u601d\u8def',
)

QUESTION_TYPE_BY_SHEET = {
    '\u6570\u91cf': '\u6570\u91cf\u5173\u7cfb',
    '\u6570\u63a8': '\u6570\u91cf\u5173\u7cfb',
    '\u8d44\u6599': '\u8d44\u6599\u5206\u6790',
    '\u5224\u65ad-\u56fe\u63a8': '\u5224\u65ad\u63a8\u7406',
    '\u5224\u65ad': '\u5224\u65ad\u63a8\u7406',
    '\u5224\u65ad\u539f\u56e0\u89e3\u91ca': '\u5224\u65ad\u63a8\u7406',
    '\u52a0\u5f3a-\u524a\u5f31': '\u5224\u65ad\u63a8\u7406',
    SHEET_YANYU: TYPE_YANYU,
    '\u5e38\u8bc6': '\u5e38\u8bc6\u5224\u65ad',
    '\u79d1\u63a8': '\u5224\u65ad\u63a8\u7406',
    '\u7c7b\u6bd4\u5b9a\u4e49': '\u5224\u65ad\u63a8\u7406',
    '\u5b9a\u4e49': '\u5224\u65ad\u63a8\u7406',
}

WORKBOOK_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
DRAWING_NS = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
MAIN_DRAWING_NS = 'http://schemas.openxmlformats.org/drawingml/2006/main'


@dataclass
class ReviewCandidate:
    sheet: str
    row_index: int
    reason: str
    raw: Dict[str, str]


@dataclass
class ParsedRow:
    sheet: str
    row_index: int
    type: str
    subtype: str
    sub_subtype: str
    question: str
    answer: str
    error_reason: str
    root_reason: str
    analysis: str
    review_status: str
    paper_ref: str
    similar_questions: str
    img_data: str | None = None
    analysis_img_data: str | None = None


@dataclass
class SheetImage:
    row_index: int
    col_index: int
    data_url: str
    source_path: str


def norm(value: object) -> str:
    text = '' if value is None else str(value)
    text = text.replace('\u3000', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def join_non_empty(parts: Iterable[str], sep: str = '\n') -> str:
    return sep.join([norm(part) for part in parts if norm(part)])


class WorkbookConverter:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.question_rows: List[ParsedRow] = []
        self.note_blocks: DefaultDict[str, List[Tuple[str, str, str]]] = defaultdict(list)
        self.review_candidates: List[ReviewCandidate] = []
        self.sheet_images = self._load_sheet_images()
        self.sheet_parsers: Dict[str, Callable[[str, List[Sequence[object]]], None]] = {
            SHEET_YANYU: self._parse_yanyu_sheet,
        }

    def run(self) -> Dict[str, object]:
        for sheet_name, rows in self._load_workbook_rows():
            self._parse_sheet(sheet_name, rows)

        return {
            'xc_version': 2,
            'exportTime': datetime.now().isoformat(timespec='seconds'),
            'errors': [self._build_error_payload(row, index + 1) for index, row in enumerate(self.question_rows)],
            'notesByType': self._build_notes_payload(),
            'noteImages': {},
            'typeRules': None,
            'dirTree': None,
            'globalNote': f'Imported from {self.workbook_path.name} on {date.today().isoformat()}',
        }

    def _parse_sheet(self, sheet_name: str, rows: List[Sequence[object]]) -> None:
        if not rows:
            return

        parser = self.sheet_parsers.get(sheet_name)
        if parser:
            parser(sheet_name, rows)
            return

        self._mark_sheet_pending(sheet_name, rows)

    def _mark_sheet_pending(self, sheet_name: str, rows: List[Sequence[object]]) -> None:
        headers = self._read_headers(rows[0]) if rows else []
        for row_index, row in enumerate(rows[1:], start=2):
            raw = self._row_to_dict(headers, row) if headers else {'raw': join_non_empty(row, ' | ')}
            if not any(raw.values()):
                continue
            self.review_candidates.append(ReviewCandidate(sheet_name, row_index, 'sheet_rule_pending', raw))

    def _load_workbook_rows(self) -> List[Tuple[str, List[Sequence[object]]]]:
        try:
            wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
            return [(sheet.title, list(sheet.iter_rows(values_only=True))) for sheet in wb.worksheets]
        except Exception as exc:
            return self._load_via_node_fallback(exc)

    def _load_via_node_fallback(self, original_exc: Exception) -> List[Tuple[str, List[Sequence[object]]]]:
        helper = Path(__file__).resolve().parents[1] / 'tools' / 'dump_workbook.js'
        command = ['node', str(helper), str(self.workbook_path)]
        completed = subprocess.run(command, capture_output=True, check=False)
        if completed.returncode != 0:
            stderr = completed.stderr.decode('utf-8', errors='replace').strip()
            raise RuntimeError(f'Workbook read failed with openpyxl ({original_exc}) and node fallback ({stderr})')
        payload = json.loads(completed.stdout.decode('utf-8', errors='replace'))
        return [(item['name'], item['rows']) for item in payload]

    def _parse_yanyu_sheet(self, sheet_name: str, rows: List[Sequence[object]]) -> None:
        headers = self._read_headers(rows[0])
        if not headers:
            return

        last_subtype = UNCATEGORIZED
        for row_index, row in enumerate(rows[1:], start=2):
            data = self._row_to_dict(headers, row)
            if not any(data.values()):
                continue

            subtype = norm(data.get('type')) or last_subtype
            question = norm(data.get('question'))
            answer = norm(data.get('answer'))
            error_reason = norm(data.get('error_reason'))
            root_reason = norm(data.get('root_reason'))
            analysis = norm(data.get('analysis'))
            review_status = norm(data.get('review_status'))
            paper_ref = norm(data.get('paper_ref'))
            similar_questions = norm(data.get('similar_questions'))

            if subtype:
                last_subtype = subtype

            row_images = self._get_row_images(sheet_name, row_index)
            question_image = self._pick_question_image(row_images)
            analysis_image = self._pick_analysis_image(row_images, question_image)

            if self._is_yanyu_note_row(question, answer, error_reason, root_reason, analysis, question_image):
                title = question or error_reason or subtype or f'{sheet_name}-{row_index}'
                body = join_non_empty(
                    [
                        analysis,
                        root_reason,
                        error_reason,
                        review_status,
                        similar_questions,
                        f'![{IMAGE_PROMPT_FALLBACK}]({question_image.data_url})' if question_image else '',
                    ],
                    '\n\n',
                )
                self.note_blocks[TYPE_YANYU].append((subtype, title, body))
                continue

            if self._is_yanyu_question_row(answer, question, question_image):
                prompt = question or IMAGE_PROMPT_FALLBACK
                self.question_rows.append(
                    ParsedRow(
                        sheet=sheet_name,
                        row_index=row_index,
                        type=TYPE_YANYU,
                        subtype=subtype or UNCATEGORIZED,
                        sub_subtype='',
                        question=prompt,
                        answer=answer,
                        error_reason=error_reason,
                        root_reason=root_reason,
                        analysis=analysis,
                        review_status=review_status,
                        paper_ref=paper_ref,
                        similar_questions=similar_questions,
                        img_data=question_image.data_url if question_image else None,
                        analysis_img_data=analysis_image.data_url if analysis_image else None,
                    )
                )
                continue

            reason = 'ambiguous_row'
            if question_image and not answer:
                reason = 'image_row_without_answer'
            elif answer and not self._looks_like_choice_answer(answer):
                reason = 'non_choice_answer'
            elif question and not answer:
                reason = 'question_without_answer'
            self.review_candidates.append(ReviewCandidate(sheet_name, row_index, reason, data))

    def _read_headers(self, first_row: Sequence[object]) -> List[str]:
        headers: List[str] = []
        for index, cell in enumerate(first_row):
            raw = norm(cell)
            headers.append(QUESTION_HEADER_ALIASES.get(raw, f'col_{index}'))
        return headers

    def _row_to_dict(self, headers: List[str], row: Sequence[object]) -> Dict[str, str]:
        result: Dict[str, str] = {}
        for index, header in enumerate(headers):
            result[header] = norm(row[index] if index < len(row) else '')
        return result

    def _is_yanyu_note_row(
        self,
        question: str,
        answer: str,
        error_reason: str,
        root_reason: str,
        analysis: str,
        question_image: SheetImage | None,
    ) -> bool:
        if question and any(keyword in question for keyword in ('\u5982\u4f55', '\u603b\u7ed3', '\u600e\u4e48\u529e', '\u7ea0\u9519')):
            return True
        if question.startswith('\u6982\u62ec\u603b\u7ed3'):
            return True
        if not self._looks_like_choice_answer(answer) and (analysis or root_reason or error_reason):
            return True
        if not answer and not question_image and (analysis or root_reason or error_reason):
            return True
        return False

    def _is_yanyu_question_row(self, answer: str, question: str, question_image: SheetImage | None) -> bool:
        if not self._looks_like_choice_answer(answer):
            return False
        return bool(question or question_image)

    def _looks_like_choice_answer(self, answer: str) -> bool:
        return bool(re.fullmatch(r'[A-D](?:[、,/][A-D])*', answer))

    def _get_row_images(self, sheet_name: str, row_index: int) -> List[SheetImage]:
        return sorted(self.sheet_images.get(sheet_name, {}).get(row_index, []), key=lambda item: item.col_index)

    def _pick_question_image(self, images: List[SheetImage]) -> SheetImage | None:
        for image in images:
            if image.col_index <= 3:
                return image
        return images[0] if images else None

    def _pick_analysis_image(self, images: List[SheetImage], question_image: SheetImage | None) -> SheetImage | None:
        for image in images:
            if question_image and image.source_path == question_image.source_path:
                continue
            if image.col_index >= 4:
                return image
        for image in images:
            if question_image and image.source_path == question_image.source_path:
                continue
            return image
        return None

    def _load_sheet_images(self) -> DefaultDict[str, DefaultDict[int, List[SheetImage]]]:
        grouped: DefaultDict[str, DefaultDict[int, List[SheetImage]]] = defaultdict(lambda: defaultdict(list))
        try:
            with zipfile.ZipFile(self.workbook_path) as archive:
                names = set(archive.namelist())
                workbook = ET.fromstring(archive.read('xl/workbook.xml'))
                workbook_rels = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
                workbook_rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in workbook_rels}

                for sheet in workbook.findall(f'{{{WORKBOOK_NS}}}sheets/{{{WORKBOOK_NS}}}sheet'):
                    sheet_name = sheet.attrib['name']
                    rel_id = sheet.attrib[f'{{{REL_NS}}}id']
                    sheet_target = workbook_rel_map.get(rel_id)
                    if not sheet_target:
                        continue

                    sheet_path = self._resolve_zip_target('xl/workbook.xml', sheet_target)
                    rels_path = posixpath.join(posixpath.dirname(sheet_path), '_rels', posixpath.basename(sheet_path) + '.rels')
                    if rels_path not in names:
                        continue

                    sheet_rels = ET.fromstring(archive.read(rels_path))
                    drawing_target = None
                    for rel in sheet_rels:
                        if rel.attrib.get('Type', '').endswith('/drawing'):
                            drawing_target = self._resolve_zip_target(sheet_path, rel.attrib['Target'])
                            break
                    if not drawing_target:
                        continue

                    drawing_rels_path = posixpath.join(
                        posixpath.dirname(drawing_target),
                        '_rels',
                        posixpath.basename(drawing_target) + '.rels',
                    )
                    if drawing_rels_path not in names:
                        continue

                    drawing_rels = ET.fromstring(archive.read(drawing_rels_path))
                    drawing_rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in drawing_rels}
                    drawing = ET.fromstring(archive.read(drawing_target))

                    for anchor in list(drawing):
                        marker = anchor.find(f'{{{DRAWING_NS}}}from')
                        blip = anchor.find(f'.//{{{MAIN_DRAWING_NS}}}blip')
                        if marker is None or blip is None:
                            continue

                        image_rel_id = blip.attrib.get(f'{{{REL_NS}}}embed')
                        image_target = drawing_rel_map.get(image_rel_id)
                        if not image_target:
                            continue

                        row_index = int(marker.findtext(f'{{{DRAWING_NS}}}row', default='0')) + 1
                        col_index = int(marker.findtext(f'{{{DRAWING_NS}}}col', default='0')) + 1
                        image_path = self._resolve_zip_target(drawing_target, image_target)
                        image_bytes = archive.read(image_path)
                        grouped[sheet_name][row_index].append(
                            SheetImage(
                                row_index=row_index,
                                col_index=col_index,
                                data_url=self._to_data_url(image_path, image_bytes),
                                source_path=image_path,
                            )
                        )
        except Exception:
            return grouped
        return grouped

    def _resolve_zip_target(self, source_path: str, target: str) -> str:
        if target.startswith('/'):
            return target.lstrip('/')
        return posixpath.normpath(posixpath.join(posixpath.dirname(source_path), target))

    def _to_data_url(self, image_path: str, image_bytes: bytes) -> str:
        mime = {
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.gif': 'image/gif',
            '.bmp': 'image/bmp',
        }.get(Path(image_path).suffix.lower(), 'application/octet-stream')
        return f'data:{mime};base64,{base64.b64encode(image_bytes).decode("ascii")}'

    def _build_error_payload(self, row: ParsedRow, item_id: int) -> Dict[str, object]:
        analysis_parts = [
            row.analysis,
            row.root_reason,
            f'\u539f\u59cb\u9898\u53f7/\u5957\u5377: {row.paper_ref}' if row.paper_ref else '',
            f'\u7c7b\u4f3c\u9898\u76ee: {row.similar_questions}' if row.similar_questions else '',
            row.review_status,
        ]
        return {
            'id': item_id,
            'addDate': date.today().isoformat(),
            'type': row.type or UNCATEGORIZED,
            'subtype': row.subtype or UNCATEGORIZED,
            'subSubtype': row.sub_subtype or '',
            'question': row.question,
            'options': '',
            'answer': row.answer,
            'myAnswer': '',
            'rootReason': row.root_reason,
            'errorReason': row.error_reason,
            'analysis': join_non_empty(analysis_parts, '\n\n'),
            'status': 'focus',
            'difficulty': 0,
            'imgData': row.img_data,
            'analysisImgData': row.analysis_img_data,
            'srcYear': '',
            'srcProvince': '',
            'srcOrigin': f'{self.workbook_path.name}::{row.sheet}',
            'quiz': None,
        }

    def _build_notes_payload(self) -> Dict[str, object]:
        notes: Dict[str, object] = {}
        for note_type, entries in sorted(self.note_blocks.items()):
            grouped: DefaultDict[str, List[Tuple[str, str]]] = defaultdict(list)
            for subtype, title, body in entries:
                grouped[subtype or UNCATEGORIZED].append((title, body))

            sections: List[str] = []
            for subtype, blocks in grouped.items():
                sections.append(f'## {subtype}')
                for title, body in blocks:
                    sections.append(f'### {title}')
                    if body:
                        sections.append(body)
                sections.append('')

            notes[note_type] = {
                'content': '\n\n'.join(sections).strip(),
                'updatedAt': datetime.now().isoformat(timespec='seconds'),
                'children': {},
            }
        return notes


def write_outputs(output_dir: Path, backup: Dict[str, object], review_candidates: List[ReviewCandidate]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / 'xingce_backup.json').write_text(json.dumps(backup, ensure_ascii=False, indent=2), encoding='utf-8')
    (output_dir / 'review_candidates.json').write_text(
        json.dumps([asdict(item) for item in review_candidates], ensure_ascii=False, indent=2),
        encoding='utf-8',
    )

    notes_preview = []
    for note_type, payload in (backup.get('notesByType') or {}).items():
        content = payload.get('content', '') if isinstance(payload, dict) else ''
        notes_preview.append(f'# {note_type}\n\n{content}\n')
    (output_dir / 'notes_preview.md').write_text('\n\n'.join(notes_preview), encoding='utf-8')

    summary = {
        'error_count': len(backup.get('errors', [])),
        'note_type_count': len(backup.get('notesByType', {})),
        'review_candidate_count': len(review_candidates),
    }
    (output_dir / 'summary.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')


def main() -> None:
    parser = argparse.ArgumentParser(description='Convert workbook rows into xingce_v3 backup JSON.')
    parser.add_argument('--input', required=True, help='Path to the source xlsx workbook.')
    parser.add_argument('--output-dir', default=str(Path(__file__).resolve().parents[1] / 'output'))
    args = parser.parse_args()

    workbook_path = Path(args.input)
    if not workbook_path.exists():
        raise SystemExit(f'Input not found: {workbook_path}')

    converter = WorkbookConverter(workbook_path)
    backup = converter.run()
    write_outputs(Path(args.output_dir), backup, converter.review_candidates)

    print(f'Generated {len(backup["errors"])} errors')
    print(f'Generated {len(backup["notesByType"])} note buckets')
    print(f'Generated {len(converter.review_candidates)} review candidates')


if __name__ == '__main__':
    main()
